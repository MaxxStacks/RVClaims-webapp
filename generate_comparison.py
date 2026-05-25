"""
V6 vs V7 Sidebar Navigation Comparison Generator
DealerSuite360 — All Roles
Reads from actual portal/layout files (already done manually above).
"""

try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'V6-V7-ROLE-COMPARISON.xlsx')

# ─── Color constants ──────────────────────────────────────────────────────────
HDR_BG   = "033280"   # navy
HDR_FG   = "FFFFFF"
GREEN    = "C6EFCE"
YELLOW   = "FFEB9C"
ORANGE   = "FCD5B4"
LRED     = "FFC7CE"
RED_BG   = "FF0000"
RED_FG   = "FFFFFF"
PURPLE_C = "CCC0DA"

TAB_PURPLE = "7030A0"
TAB_GREEN  = "375623"
TAB_BLUE   = "1F4E79"
TAB_ORANGE = "C55A11"
TAB_GRAY   = "595959"

STATUS_COLORS = {
    "LIVE":        (GREEN,   "000000"),
    "PARTIAL":     (YELLOW,  "000000"),
    "STUB":        (ORANGE,  "000000"),
    "PLACEHOLDER": (LRED,    "000000"),
    "MISSING":     (RED_BG,  RED_FG),
    "V6 ONLY":     (PURPLE_C,"000000"),
    "N/A":         ("F2F2F2","888888"),
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style='thin', color='D0D0D0')
    return Border(left=s, right=s, top=s, bottom=s)

def write_sheet(wb, sheet_name, tab_color_hex, headers, rows):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color_hex

    # Header row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = fill(HDR_BG)
        cell.font = Font(bold=True, color=HDR_FG, size=11, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Column widths
    col_widths = [45, 45, 15, 60, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        v6_item, v7_item, status, file_path, isolatable = row_data
        vals = [v6_item, v7_item, status, file_path, isolatable]

        status_bg, status_fg = STATUS_COLORS.get(status, ("FFFFFF", "000000"))

        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in (1,2,4)))
            cell.border = thin_border()
            cell.font = Font(name="Calibri", size=10)

            # Status column coloring
            if col_idx == 3:
                cell.fill = fill(status_bg)
                cell.font = Font(name="Calibri", size=10, bold=True, color=status_fg)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Shade V6-only rows lightly
            if not v6_item and col_idx != 3:
                cell.fill = fill("F8F8FF")
            if not v7_item and col_idx != 3:
                cell.fill = fill("FFF8F0")

    return ws

# ─── PAGE FILES — presence check ─────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
PAGES_DIR = os.path.join(BASE, "client", "src", "pages")

def page_exists(filename):
    """Return True if the page file exists in client/src/pages/"""
    p = os.path.join(PAGES_DIR, filename)
    return os.path.isfile(p)

def page_status(filename):
    """
    Read page file and determine status:
    LIVE     = has real API calls (apiFetch, fetch(, axios, useQuery)
    PARTIAL  = mix of real API + hardcoded/alert/console.log
    STUB     = hardcoded data, no real API calls
    PLACEHOLDER = only a heading or coming soon text
    MISSING  = file not found
    """
    p = os.path.join(PAGES_DIR, filename)
    if not os.path.isfile(p):
        return "MISSING"
    with open(p, encoding="utf-8", errors="ignore") as f:
        content = f.read()

    has_api = any(x in content for x in [
        "apiFetch", "fetch(", "axios", "useQuery", "useMutation",
        "/api/", "tanstack"
    ])
    has_stub = any(x in content for x in [
        "alert(", "console.log", "hardcoded", "TODO", "Coming soon",
        "coming soon", "placeholder", "lorem"
    ])
    has_content = len(content.strip()) > 200

    if not has_content:
        return "PLACEHOLDER"
    if has_api and has_stub:
        return "PARTIAL"
    if has_api:
        return "LIVE"
    if has_stub:
        return "STUB"
    return "STUB"

def p(filename, v6_label=None):
    """Return (status, file_path) tuple for a page file."""
    status = page_status(filename)
    rel_path = f"client/src/pages/{filename}" if filename else ""
    return status, rel_path

# ─── DATA DEFINITIONS ─────────────────────────────────────────────────────────
# Format: (v6_item, v7_item, status, file_path, isolatable)
# v6_item / v7_item = "SECTION | Item Name"
# Empty string = not present in that version

def make_rows_op_admin():
    """Operator Admin — V6 OperatorPortal (Admin view) vs V7 OperatorAdminLayout"""
    # V6 has both Admin+Staff in one portal; V7 splits into OperatorAdminLayout / OperatorStaffLayout
    # Pages from pages/ dir + status checks
    rows = [
        # Overview
        ("Overview | Dashboard",        "Overview | Dashboard",       *p("Dashboard.tsx"),          "YES"),
        # Claims
        ("Claims | Processing Queue",   "Claims | Processing Queue",  *p("ProcessingQueue.tsx"),    "YES"),
        ("Claims | All Claims",         "Claims | Claims",            *p("Claims.tsx"),              "YES"),
        ("Claims | Stale Claims",       "Claims | Stale Claims",      *p("StaleClaims.tsx"),         "YES"),
        ("Claims | Batch Review",       "",                           "V6 ONLY", "client/src/pages/BatchReview.tsx", "YES (with deps)"),
        # Management
        ("Management | Dealers",        "Management | Dealers",       *p("DealerManagement.tsx"),   "YES"),
        ("Management | Units",          "Management | Units",         *p("Units.tsx"),              "YES"),
        ("Management | FRC Codes",      "Management | FRC Codes",     *p("FRCCodes.tsx"),           "YES"),
        ("Management | Mfr Portals",    "",                           "V6 ONLY", "client/src/pages/MfrPortals.tsx", "PARTIAL"),
        ("Management | TechFlow Oversight", "",                       "V6 ONLY", "",               "N/A"),
        ("Management | Work by Dealer *","",                          "V6 ONLY", "",               "N/A"),
        ("Management | Campaign Templates *","",                      "V6 ONLY", "",               "N/A"),
        ("Management | Consignment Oversight *","",                   "V6 ONLY", "",               "N/A"),
        # Services
        ("Services | Service Marketplace","Services | Marketplace",   *p("ServiceMarketplace.tsx"), "YES"),
        ("Services | Financing",        "Services | Financing",       *p("Financing.tsx"),          "YES"),
        ("Services | Financing Apps",   "",                           "V6 ONLY", "",               "N/A"),
        ("Services | Parts Catalog",    "",                           "V6 ONLY", "",               "N/A"),
        ("Services | F&I Services",     "Services | F&I",            *p("FAndI.tsx"),              "YES"),
        ("Services | Warranty Plans",   "Services | Warranty Plans",  *p("WarrantyPlans.tsx"),      "YES"),
        ("Services | Parts & Accessories","Services | Parts",         *p("Parts.tsx"),              "YES"),
        ("Services | Parts Management *","",                          "V6 ONLY", "",               "N/A"),
        ("Services | Parts Orders *",   "",                           "V6 ONLY", "",               "N/A"),
        # Marketplace
        ("Marketplace | Members",       "",                           "V6 ONLY", "",               "N/A"),
        ("Marketplace | Listings",      "",                           "V6 ONLY", "",               "N/A"),
        ("Marketplace | Transactions",  "",                           "V6 ONLY", "",               "N/A"),
        ("Marketplace | Live Auctions", "",                           "V6 ONLY", "",               "N/A"),
        ("Marketplace | Public Auctions","",                          "V6 ONLY", "",               "N/A"),
        ("Marketplace | Escrow Admin *","",                           "V6 ONLY", "",               "N/A"),
        # Finance (Admin only in V6)
        ("Finance | Billing & Invoices","Finance | Invoices",         *p("Invoices.tsx"),           "YES"),
        ("Finance | Products & Services","Finance | Products",        *p("Products.tsx"),           "YES"),
        ("Finance | Revenue Reports",   "Finance | Reports",          *p("Reports.tsx"),            "YES"),
        ("Finance | Financing Partners","",                           "V6 ONLY", "",               "N/A"),
        # CRM
        ("CRM | Dealer Directory",      "CRM | CRM",                 *p("Dashboard.tsx"),           "YES (with deps)"),
        # System
        ("System | Notifications",      "System | Send Notifications",*p("Notifications.tsx"),      "YES"),
        ("System | Communications",     "System | Communications",   "",                            "MISSING", "NO"),
        ("System | Users & Roles",      "System | Users & Roles",    *p("UsersRoles.tsx"),          "YES"),
        ("System | Settings",           "System | Settings",         *p("Settings.tsx"),            "YES"),
        ("System | Changelog",          "System | Changelog",        *p("Changelog.tsx"),           "YES"),
        ("System | Blog",               "System | Blog",             "",                            "MISSING", "NO"),
        ("System | System Status",      "",                           "V6 ONLY", "",               "N/A"),
        # V7 only
        ("",                            "System | Platform Settings", "MISSING", "",               "N/A"),
        ("",                            "System | Import Data",       *p("ImportData.tsx"),          "N/A"),
        ("",                            "System | Roadmap",           "",                            "MISSING", "N/A"),
    ]
    return rows

def make_rows_op_staff():
    """Operator Staff — V6 OperatorPortal (Staff view) vs V7 OperatorStaffLayout"""
    rows = [
        # Overview
        ("Overview | Dashboard",        "Overview | Dashboard",       *p("Dashboard.tsx"),          "YES"),
        # Claims
        ("Claims | Processing Queue",   "Claims | Processing Queue",  *p("ProcessingQueue.tsx"),    "YES"),
        ("Claims | All Claims",         "Claims | Claims",            *p("Claims.tsx"),             "YES"),
        ("Claims | Stale Claims",       "Claims | Stale Claims",      *p("StaleClaims.tsx"),        "YES"),
        # Management
        ("Management | Dealers",        "Management | Dealers",       *p("DealerManagement.tsx"),   "YES"),
        ("Management | Units",          "Management | Units",         *p("Units.tsx"),              "YES"),
        ("Management | FRC Codes",      "",                           "V6 ONLY", "client/src/pages/FRCCodes.tsx", "YES"),
        ("",                            "Management | Parts",         *p("Parts.tsx"),              "YES"),
        # Services (Staff had all Services items in V6 too)
        ("Services | Service Marketplace","",                         "V6 ONLY", "client/src/pages/ServiceMarketplace.tsx", "YES"),
        ("Services | Financing",        "",                           "V6 ONLY", "client/src/pages/Financing.tsx", "YES"),
        ("Services | F&I Services",     "",                           "V6 ONLY", "client/src/pages/FAndI.tsx", "YES"),
        ("Services | Warranty Plans",   "",                           "V6 ONLY", "client/src/pages/WarrantyPlans.tsx", "YES"),
        ("Services | Parts & Accessories","",                         "V6 ONLY", "client/src/pages/Parts.tsx", "YES"),
        # Marketplace (V6 had these for all operators)
        ("Marketplace | Members",       "",                           "V6 ONLY", "",               "N/A"),
        ("Marketplace | Listings",      "",                           "V6 ONLY", "",               "N/A"),
        ("Marketplace | Transactions",  "",                           "V6 ONLY", "",               "N/A"),
        ("Marketplace | Live Auctions", "",                           "V6 ONLY", "",               "N/A"),
        ("Marketplace | Public Auctions","",                          "V6 ONLY", "",               "N/A"),
        # CRM
        ("CRM | Dealer Directory",      "",                           "V6 ONLY", "",               "N/A"),
        # System
        ("System | Notifications",      "System | Notifications",    *p("Notifications.tsx"),      "YES"),
        ("System | Communications",     "",                           "V6 ONLY", "",               "N/A"),
        ("System | Changelog",          "System | Changelog",        *p("Changelog.tsx"),          "YES"),
        ("System | Blog",               "",                           "V6 ONLY", "",               "N/A"),
        ("System | System Status",      "",                           "V6 ONLY", "",               "N/A"),
    ]
    return rows

def make_rows_dealer_owner():
    """Dealer Owner — V6 DealerPortal (isDealerOwner) vs V7 DealerOwnerLayout"""
    rows = [
        # Overview
        ("Overview | Dashboard",        "Overview | Dashboard",       *p("Dashboard.tsx"),          "YES"),
        # Claims
        ("Claims | Upload Photos",      "Claims | Upload Photos",     *p("ClaimNew.tsx"),           "YES (with deps)"),
        ("Claims | My Claims",          "Claims | Claims",            *p("Claims.tsx"),             "YES"),
        ("Units | My Units",            "Claims | Units",             *p("Units.tsx"),              "YES"),
        # Services
        ("Services | Financing",        "Services | Financing",       *p("Financing.tsx"),          "YES"),
        ("Services | F&I Products",     "Services | F&I",            *p("FAndI.tsx"),              "YES"),
        ("Services | Warranty Plans",   "Services | Warranty Plans",  *p("Warranty.tsx"),           "YES"),
        ("Services | Parts Orders",     "Services | Parts",          *p("Parts.tsx"),              "YES"),
        # Marketplace
        ("Marketplace | Browse Units",  "Market | Marketplace",      *p("Auctions.tsx"),           "STUB"),
        ("Marketplace | My Listings",   "",                           "V6 ONLY", "",               "N/A"),
        ("Marketplace | My Transactions","",                          "V6 ONLY", "",               "N/A"),
        ("Marketplace | Live Auctions", "",                           "V6 ONLY", "",               "N/A"),
        ("Marketplace | Public Showcase","",                          "V6 ONLY", "",               "N/A"),
        ("Marketplace | My Bids",       "",                           "V6 ONLY", "",               "N/A"),
        ("Marketplace | Escrow & Payments","",                        "V6 ONLY", "",               "N/A"),
        # Operations
        ("Operations | Client Files",   "Customers | Customers",     *p("Customers.tsx"),          "YES"),
        ("Operations | Messages",       "Customers | Messages",      *p("Messages.tsx"),           "PARTIAL"),
        ("Operations | Documents",      "Customers | Documents",     *p("Documents.tsx"),          "STUB"),
        ("Operations | TechFlow",       "Customers | TechFlow",      *p("WorkOrders.tsx"),         "PARTIAL"),
        ("Operations | Consignment",    "Market | Consignment",      "",                           "MISSING", "NO"),
        ("Operations | Marketing",      "Market | Marketing",        "",                           "MISSING", "NO"),
        ("Operations | Sales & Services","Market | Sales & Services","",                           "MISSING", "NO"),
        # Customers
        ("Customers | Customer Portal", "Customers | Customer Tickets","",                         "STUB", "YES (with deps)"),
        ("Customers | Customer Tickets","",                           "V6 ONLY", "client/src/pages/CustomerTickets.tsx", "YES"),
        # Billing
        ("Billing | Invoices & Billing","Billing | Invoices",        *p("Invoices.tsx"),           "YES"),
        ("Billing | My Subscription",   "Billing | Billing",         "",                           "MISSING", "NO"),
        ("Billing | Portal Settings",   "",                           "V6 ONLY", "",               "N/A"),
        # Settings
        ("Settings | Staff",            "Settings | Staff",          "",                           "MISSING", "NO"),
        ("Settings | Branding & Domain","Settings | Branding",       "",                           "STUB", "YES"),
        ("Settings | Settings",         "Settings | Settings",       *p("Settings.tsx"),           "PARTIAL"),
        ("Settings | Notifications",    "Settings | Notifications",  *p("Notifications.tsx"),      "YES"),
        ("Settings | What's New",       "Settings | What's New",     *p("Changelog.tsx"),          "PARTIAL"),
        ("Settings | System Status",    "",                           "V6 ONLY", "",               "N/A"),
        # V7 only
        ("",                            "Customers | Documents",     *p("Documents.tsx"),          "N/A"),
        ("",                            "Settings | Import Data",    *p("ImportData.tsx"),          "N/A"),
    ]
    return rows

def make_rows_dealer_staff():
    """Dealer Staff — V6 DealerPortal (isDealerStaff) vs V7 DealerStaffLayout"""
    rows = [
        ("Overview | Dashboard",        "Overview | Dashboard",       *p("Dashboard.tsx"),         "YES"),
        # Claims
        ("Claims | Upload Photos",      "Claims | Upload Photos",     *p("ClaimNew.tsx"),          "YES (with deps)"),
        ("Claims | My Claims",          "Claims | Claims",            *p("Claims.tsx"),            "YES"),
        ("Units | My Units",            "Claims | Units",             *p("Units.tsx"),             "YES"),
        ("",                            "Claims | Parts",             *p("Parts.tsx"),             "YES"),
        # Customers
        ("Customers | Customer Portal", "Customers | Customers",     *p("Customers.tsx"),         "YES"),
        ("Customers | Customer Tickets","Customers | Customer Tickets","",                         "STUB", "YES (with deps)"),
        ("Operations | Documents",      "Customers | Documents",     *p("Documents.tsx"),         "STUB"),
        ("Operations | Messages",       "Customers | Messages",      *p("Messages.tsx"),          "PARTIAL"),
        # System
        ("Settings | Notifications",    "System | Notifications",    *p("Notifications.tsx"),     "YES"),
        ("Settings | What's New",       "System | What's New",       *p("Changelog.tsx"),         "PARTIAL"),
        ("Settings | System Status",    "",                           "V6 ONLY", "",              "N/A"),
        ("",                            "System | Import Data",      *p("ImportData.tsx"),         "N/A"),
        # V6 items staff could access but V7 removed
        ("Services | Parts Orders",     "",                           "V6 ONLY", "client/src/pages/Parts.tsx", "YES"),
        ("Operations | TechFlow",       "",                           "V6 ONLY", "client/src/pages/WorkOrders.tsx", "PARTIAL"),
        ("Operations | Messages",       "",                           "V6 ONLY", "client/src/pages/Messages.tsx", "PARTIAL"),
    ]
    return rows

def make_rows_technician():
    """Technician (On-Site Tech) — V6 DealerPortal (isTechnician) vs V7 OnSiteTechLayout"""
    rows = [
        ("Overview | Dashboard",        "Overview | Dashboard",       *p("Dashboard.tsx"),         "YES"),
        ("Units | My Units",            "",                           "V6 ONLY", "client/src/pages/Units.tsx", "YES"),
        ("Services | Parts Orders",     "My Work | Parts Request",   *p("Parts.tsx"),             "YES"),
        ("Operations | TechFlow",       "My Work | My Work Orders",  *p("WorkOrders.tsx"),         "PARTIAL"),
        ("Settings | Notifications",    "",                           "V6 ONLY", "client/src/pages/Notifications.tsx", "YES"),
        ("Settings | Settings",         "",                           "V6 ONLY", "client/src/pages/Settings.tsx", "PARTIAL"),
        ("Settings | What's New",       "",                           "V6 ONLY", "client/src/pages/Changelog.tsx", "PARTIAL"),
        # V7 only items
        ("",                            "My Work | My Route",        "MISSING", "",               "N/A"),
        ("",                            "My Work | Time Clock",      "MISSING", "",               "N/A"),
    ]
    return rows

def make_rows_client():
    """Client (Customer) — V6 CustomerPortal vs V7 ClientLayout"""
    rows = [
        ("Overview | Dashboard",        "Overview | Dashboard",       *p("Dashboard.tsx"),         "YES"),
        # My RV
        ("My RV | My Unit",             "My RV | My Unit",           "",                          "STUB", "YES (with deps)"),
        ("My RV | Warranty & Coverage", "My RV | Warranty",          *p("Warranty.tsx"),          "PARTIAL"),
        ("My RV | Documents",           "My RV | Documents",         *p("Documents.tsx"),         "STUB"),
        # Service
        ("Service | Claim Status",      "My RV | Claims",            *p("Claims.tsx"),            "PARTIAL"),
        ("Service | Report an Issue",   "",                           "V6 ONLY", "client/src/pages/ClaimNew.tsx", "YES (with deps)"),
        ("Service | Parts Orders",      "Services | Parts",          *p("Parts.tsx"),             "PARTIAL"),
        # Extras / Services
        ("Extras | Protection Plans",   "Services | F&I Products",   *p("FAndI.tsx"),             "STUB"),
        ("Extras | My Offers",          "Services | F&I Offers",     "",                          "MISSING", "NO"),
        ("Extras | My Financing",       "Services | My Financing",   *p("Financing.tsx"),         "PARTIAL"),
        ("Extras | Appointments",       "Services | Service Appointments","",                     "STUB", "YES (with deps)"),
        ("Extras | Roadside Assist",    "Services | Roadside",       "",                          "PLACEHOLDER", "NO"),
        ("Extras | My Services *",      "Services | My Services",    "",                          "MISSING", "NO"),
        # Account
        ("Account | Messages",          "Account | Messages",        *p("Messages.tsx"),          "PARTIAL"),
        ("Account | Support Tickets",   "Account | Support Tickets", "",                          "STUB", "YES (with deps)"),
        ("Account | Quick Chat",        "Account | Quick Chat",      "",                          "STUB", "YES (with deps)"),
        ("Account | Account *",         "",                          "V6 ONLY", "",               "N/A"),
        ("Account | Settings",          "Account | Settings",        *p("Settings.tsx"),          "PARTIAL"),
        ("Account | System Status",     "",                          "V6 ONLY", "",               "N/A"),
    ]
    return rows

def make_rows_public_bidder():
    """Public Bidder — V6 BidderPortal + DealerPortal(isPublicBidder) vs V7 PublicBidderLayout"""
    rows = [
        ("Account | Dashboard",         "Account | Dashboard",        *p("Dashboard.tsx"),         "STUB"),
        ("Account | My Profile",        "Account | My Profile",      "",                          "STUB", "YES"),
        ("Account | Verification",      "Account | ID Verification", "",                          "STUB", "YES"),
        ("Account | Payment & Card",    "Account | Payment & Card",  "",                          "STUB", "YES"),
        ("Account | Payment Methods",   "Account | Payment Methods", "",                          "STUB", "YES"),
        # Auctions
        ("Auctions | Browse Units",     "Auctions | Upcoming Auctions","",                        "STUB", "YES (with deps)"),
        ("Auctions | Upcoming",         "",                           "V6 ONLY", "",              "N/A"),
        ("Auctions | My Bids",          "Auctions | My Bids",        "",                          "STUB", "YES"),
        ("Auctions | Won Units",        "Auctions | Won Units",      "",                          "STUB", "YES"),
        # Marketplace (from DealerPortal bidder view)
        ("Marketplace | Browse Units",  "",                           "V6 ONLY", "",              "N/A"),
        ("Marketplace | Live Auctions", "",                           "V6 ONLY", "",              "N/A"),
        ("My Account | My Bids",        "",                           "V6 ONLY", "",              "N/A"),
        ("My Account | Escrow & Payments","",                         "V6 ONLY", "",              "N/A"),
        # More
        ("More | Settings",             "More | Settings",           *p("Settings.tsx"),          "PARTIAL"),
        # V7 new
        ("",                            "Auctions | Upcoming Auctions","",                        "STUB", "YES"),
    ]
    return rows

def make_rows_consignor():
    """Consignor — V6 DealerPortal (isConsignor) + partial Marketplace vs V7 ConsignorLayout"""
    rows = [
        ("Overview | Dashboard",        "Overview | Dashboard",       *p("Dashboard.tsx"),        "STUB"),
        # My Consignment (V6 section)
        ("My Consignment | My Consigned Units","My Units | My Listings","",                      "STUB", "YES"),
        ("My Consignment | Offers Received","My Units | Offers",      "",                        "STUB", "YES"),
        ("My Consignment | Payouts",    "Finance | Payout Tracking", "",                         "STUB", "YES"),
        # Marketplace items consignors could see in V6
        ("Marketplace | My Listings",   "",                           "V6 ONLY", "",             "N/A"),
        ("Marketplace | Public Showcase","My Units | Showcase",       "",                        "STUB", "YES"),
        ("Marketplace | Escrow & Payments","",                        "V6 ONLY", "",             "N/A"),
        # V7 new
        ("",                            "My Units | Auctions",       "",                         "STUB", "YES"),
        ("",                            "More | Settings",           *p("Settings.tsx"),          "PARTIAL"),
    ]
    return rows

def make_rows_independent_bidder():
    """Independent Bidder — Role did not exist in V6 (handled by Public Bidder) / No V7 layout"""
    rows = [
        ("Role did not exist in V6", "Layout not implemented in V7", "MISSING", "", "N/A"),
    ]
    return rows

def make_rows_service_manager():
    """Service Manager — V6 DealerPortal (isServiceManager) vs V7 ServiceManagerLayout"""
    rows = [
        ("Overview | Dashboard",        "Overview | Dashboard",       *p("Dashboard.tsx"),        "YES"),
        # Service (V6)
        ("Service | Work Orders",       "Service | Work Orders",     *p("WorkOrders.tsx"),        "PARTIAL"),
        ("Service | Dispatch Scheduler","Service | Scheduling",      *p("WorkOrders.tsx"),        "PARTIAL"),
        ("Service | Technician Mgmt",   "Service | Technicians",     "",                          "STUB", "YES"),
        # Operations (V6)
        ("Operations | Unit Lookup",    "Service | Units",           *p("Units.tsx"),             "YES"),
        ("Operations | Parts Orders",   "Service | Parts",           *p("Parts.tsx"),             "YES"),
        ("Operations | Messages",       "",                           "V6 ONLY", "client/src/pages/Messages.tsx", "PARTIAL"),
        # V7 new
        ("",                            "Service | Claims",          *p("Claims.tsx"),            "YES"),
        ("",                            "Service | Service Appointments","",                      "STUB", "YES"),
        ("",                            "Service | Settings",        *p("Settings.tsx"),          "PARTIAL"),
    ]
    return rows

def make_rows_shop_manager():
    """Shop Manager — V6 DealerPortal (isShopManager) vs V7 ShopManagerLayout"""
    rows = [
        ("Overview | Dashboard",        "Overview | Dashboard",       *p("Dashboard.tsx"),        "YES"),
        ("Shop | Work Orders",          "Shop | Work Orders",        *p("WorkOrders.tsx"),        "PARTIAL"),
        ("Shop | Dispatch Scheduler",   "Shop | Dispatch Board",     *p("WorkOrders.tsx"),        "PARTIAL"),
        ("Shop | Parts Orders",         "Shop | Parts Status",       *p("Parts.tsx"),             "YES"),
        ("",                            "Shop | Units",              *p("Units.tsx"),             "YES"),
    ]
    return rows

def make_rows_parts_manager():
    """Parts Manager (Parts Dept) — V6 DealerPortal (isPartsDept) vs V7 PartsManagerLayout"""
    rows = [
        ("Overview | Dashboard",        "Overview | Dashboard",      *p("Dashboard.tsx"),         "YES"),
        ("Parts | Parts Orders",        "Parts | Parts Orders",      *p("Parts.tsx"),             "YES"),
        ("",                            "Parts | Inventory",         "",                          "MISSING", "NO"),
        ("",                            "Parts | Claims",            *p("Claims.tsx"),            "YES"),
    ]
    return rows

def make_rows_financial_manager():
    """Financial Manager — Role did not exist in V6 / V7 FinancialManagerLayout"""
    rows = [
        ("Role did not exist in V6",    "Overview | Dashboard",      *p("Dashboard.tsx"),         "YES"),
        ("",                            "Finance | Revenue",         *p("Reports.tsx"),           "YES"),
        ("",                            "Finance | F&I",            *p("FAndI.tsx"),              "YES"),
        ("",                            "Finance | Warranty Plans",  *p("WarrantyPlans.tsx"),     "YES"),
        ("",                            "Finance | Financing",       *p("Financing.tsx"),         "YES"),
        ("",                            "Finance | Invoices",        *p("Invoices.tsx"),          "YES"),
    ]
    return rows

def make_rows_shop_tech():
    """Shop Tech — Role did not exist in V6 / V7 ShopTechLayout"""
    rows = [
        ("Role did not exist in V6",    "Overview | Dashboard",      *p("Dashboard.tsx"),         "STUB"),
        ("",                            "My Work | My Work Orders",  *p("WorkOrders.tsx"),        "PARTIAL"),
        ("",                            "My Work | Parts Request",   *p("Parts.tsx"),             "YES"),
        ("",                            "My Work | Time Clock",      "",                          "MISSING", "NO"),
    ]
    return rows

# ─── BUILD WORKBOOK ───────────────────────────────────────────────────────────
def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    HEADERS = ["V6 Sidebar Item", "V7 Sidebar Item", "Status", "File Path", "Module Isolatable?"]

    sheets = [
        ("Op Admin",         TAB_PURPLE, make_rows_op_admin()),
        ("Op Staff",         TAB_PURPLE, make_rows_op_staff()),
        ("Dealer Owner",     TAB_GREEN,  make_rows_dealer_owner()),
        ("Dealer Staff",     TAB_GREEN,  make_rows_dealer_staff()),
        ("Technician",       TAB_GRAY,   make_rows_technician()),
        ("Client",           TAB_BLUE,   make_rows_client()),
        ("Public Bidder",    TAB_ORANGE, make_rows_public_bidder()),
        ("Consignor",        TAB_ORANGE, make_rows_consignor()),
        ("Independent Bidder", TAB_ORANGE, make_rows_independent_bidder()),
        ("Service Manager",  TAB_GRAY,   make_rows_service_manager()),
        ("Shop Manager",     TAB_GRAY,   make_rows_shop_manager()),
        ("Parts Manager",    TAB_GRAY,   make_rows_parts_manager()),
        ("Financial Manager",TAB_GRAY,   make_rows_financial_manager()),
        ("Shop Tech",        TAB_GRAY,   make_rows_shop_tech()),
    ]

    totals = {}
    for sheet_name, tab_color, rows in sheets:
        write_sheet(wb, sheet_name, tab_color, HEADERS, rows)
        totals[sheet_name] = len(rows)

    wb.save(OUTPUT_PATH)
    return totals

if __name__ == "__main__":
    print("Generating V6-V7-ROLE-COMPARISON.xlsx ...")
    totals = build_workbook()
    print(f"\nSaved to: {OUTPUT_PATH}\n")
    print("Row counts per role:")
    grand_total = 0
    for name, count in totals.items():
        print(f"  {name:<22} {count:>3} rows")
        grand_total += count
    print(f"\n  {'TOTAL':<22} {grand_total:>3} nav item comparisons")
    print("\nDone.")
